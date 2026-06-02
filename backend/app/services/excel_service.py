from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_athletes_to_excel(athletes: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Спортсмены"

    headers = ["ФИО", "Пол", "Дата рождения", "Вид спорта", "Разряд", "Статус", "Центр"]
    ws.append(headers)
    _style_header(ws)

    for a in athletes:
        ws.append([
            f"{a.get('last_name', '')} {a.get('first_name', '')} {a.get('middle_name', '') or ''}",
            {"male": "Мужской", "female": "Женский"}.get(a.get("gender", ""), a.get("gender", "")),
            a.get("birth_date", ""),
            a.get("sport_type", ""),
            a.get("rank", ""),
            a.get("status", ""),
            a.get("center_id", ""),
        ])

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_coaches_to_excel(coaches: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Тренеры"

    headers = ["ФИО", "Специализация", "Центр", "Дата найма", "Активен"]
    ws.append(headers)
    _style_header(ws)

    for c in coaches:
        ws.append([
            f"{c.get('last_name', '')} {c.get('first_name', '')} {c.get('middle_name', '') or ''}",
            c.get("specialization", ""),
            c.get("center_id", ""),
            c.get("hire_date", ""),
            "Да" if c.get("is_active") else "Нет",
        ])

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_attendance_to_excel(records: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    headers = ["Спортсмен", "Дата", "Статус", "Тренировка"]
    ws.append(headers)
    _style_header(ws)

    for r in records:
        ws.append([
            r.get("athlete_name", ""),
            str(r.get("date", "")),
            {"present": "Присутствовал", "absent": "Отсутствовал", "excused": "Уважительная"}.get(
                r.get("status", ""), r.get("status", "")
            ),
            r.get("schedule_info", ""),
        ])

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_events_to_excel(events: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "События"

    headers = ["Название", "Тип", "Дата начала", "Дата окончания", "Место", "Статус"]
    ws.append(headers)
    _style_header(ws)

    for e in events:
        ws.append([
            e.get("name", ""),
            {"tournament": "Турнир", "training_camp": "Сборы", "masterclass": "Мастер-класс"}.get(
                e.get("event_type", ""), e.get("event_type", "")
            ),
            str(e.get("start_date", "")),
            str(e.get("end_date", "")),
            e.get("location", ""),
            e.get("status", ""),
        ])

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _style_header(ws):
    fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
